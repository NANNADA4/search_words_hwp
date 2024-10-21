"""
폴더를 순회하며 스크립트를 진행합니다
"""


import os
import pathlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from natsort import natsorted
import win32com.client as win32


from module.log import write_log


def processing_folder(folder_path, excel_file, txt_path):
    """폴더를 순회하여 hwp 파일에 대해 스크립트를 진행합니다"""
    text_list = read_word(txt_path)

    for root, _, files in os.walk(folder_path):
        infos_list = []
        for file in natsorted(files):
            if not file.lower().endswith('.hwp'):
                continue
            hwp_file_path = os.path.join('\\\\?\\', root, file)
            infos_list.extend(processing_hwp(
                folder_path, hwp_file_path, text_list))
            print(f"{file} 작업 중...")

        save_infos_to_excel(infos_list, excel_file)


def processing_hwp(folder_path, hwp_file, text_list):
    """hwp 파일에서 텍스트를 찾아 결과를 반환합니다"""
    hwp_infos = []
    hwp = None

    try:
        hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
        hwp.SetMessageBoxMode(0x00000020)
        hwp.RegisterModule("FilePathCheckDLL", "SecurityModule")
        hwp.Open(hwp_file, arg="versionwarning:False;suspendpassword:True")
        hwp.InitScan()

        while True:
            state, text = hwp.GetText()
            if state in [0, 1]:
                break
            for texts in text_list:
                if texts in text:
                    hwp.MovePos(201)
                    hwp_infos.append((hwp_file, os.path.basename(hwp_file), pathlib.Path(
                        hwp_file).suffix.lstrip('.').lower(), hwp.KeyIndicator()[3],
                        texts))
                    break

    except Exception as e:  # pylint: disable=W0703
        write_log(folder_path, hwp_file, str(e))

    finally:
        if hwp:
            hwp.ReleaseScan()
            hwp.Quit()

    return hwp_infos


def read_word(txt_path):
    """텍스트 파일을 읽어 리스트로 만듭니다"""
    with open(txt_path, 'r', encoding='utf-8') as file:
        lines = [line.strip() for line in file]
    return lines


def save_infos_to_excel(infos, excel_file):
    """개인정보를 찾은 리스트를 엑셀 파일로 저장합니다."""
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = ["연번", "경로명", "파일명",
                   "확장자", "페이지번호", "내용"]
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color

    for i in range(0, len(infos), 5000):
        chunk = infos[i:i + 5000]
        for j, info in enumerate(chunk, start=ws.max_row + i):
            ws.append([j] + list(info))
        wb.save(excel_file)

    wb.save(excel_file)
